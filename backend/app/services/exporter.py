from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta


DAYS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

def export_schedule_to_excel(schedule_option, filename="schedule.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario Recomendado"

    ws.append(["Curso", "Sección", "Día", "Hora Inicio", "Hora Fin"])

    for course in schedule_option:
        for block in course["day_blocks"]:
            ws.append([
                course["course_id"],
                course["section"],
                block["day"],
                block["start"],
                block["end"]
            ])

    wb.save(filename)
    return filename


def time_range(start="07:00", end="23:00", interval=30):
    times = []
    current = datetime.strptime(start, "%H:%M")
    end_time = datetime.strptime(end, "%H:%M")

    while current < end_time:
        next_time = current + timedelta(minutes=interval)
        times.append((current.strftime("%H:%M"), next_time.strftime("%H:%M")))
        current = next_time

    return times


def export_schedule_to_excel_complex(schedule_option, filename="schedule.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"

    # Header
    ws.cell(row=1, column=1, value="Hora")
    for i, day in enumerate(DAYS, start=2):
        ws.cell(row=1, column=i, value=day)

    time_slots = time_range()

    # Crear filas de horas
    for idx, (start, end) in enumerate(time_slots, start=2):
        ws.cell(row=idx, column=1, value=f"{start}-{end}")

    # Insertar cursos
    for course in schedule_option:
        course_label = f"{course['course_id']} - {course['section']}"

        for block in course["day_blocks"]:
            day = block["day"]
            if day not in DAYS:
                continue

            col = DAYS.index(day) + 2

            for idx, (start, end) in enumerate(time_slots, start=2):
                if not (block["end"] <= start or block["start"] >= end):
                    cell = ws.cell(row=idx, column=col)
                    cell.value = course_label
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(filename)
    return filename